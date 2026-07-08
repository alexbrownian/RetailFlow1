# add_gic_design_decisions.py
# ============================
# ONE-OFF: append the GIC_RAW_DATA design decisions to design_decisions.xlsx.
# Run it ONCE on your machine (it could not be done from the assistant sandbox,
# which cannot open the .xlsx binary reliably):
#
#     python add_gic_design_decisions.py
#
# It matches your existing columns by HEADER NAME (decision, category, why,
# technical detail, outcome, date) so column order does not matter, and it
# SKIPS any decision whose 'decision' text is already present - so running it
# twice adds nothing.

import datetime
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(ROOT, "design_decisions.xlsx")
TODAY = datetime.date.today().isoformat()   # 2026-07-08 when first written

# Each decision as a plain dict keyed by the column header (lowercased).
DECISIONS = [
    {
        "decision": "Commit an abstracted aggregate layer (GIC_RAW_DATA), not raw posts",
        "category": "Data governance",
        "why": "Company rule: raw Reddit/X/StockTwits posts may not be downloaded "
               "(even via GitHub); only derived numbers may leave the raw machine. "
               "The pipeline still has to run on the work laptop.",
        "technical detail": "The five existing daily aggregates (ticker counts, "
               "counts_by_source, ticker sentiment, theme counts, theme sentiment) "
               "are published to a committed GIC_RAW_DATA/ folder via "
               "src/gic_data.py export()/hydrate(). No text, author, id or "
               "subreddit - ~2 MB total.",
        "outcome": "Work laptop runs consumer notebooks + dashboard with zero raw access.",
        "date": TODAY,
    },
    {
        "decision": "Producer/consumer split at the text->numbers line",
        "category": "Architecture",
        "why": "Notebooks 01/02/04/06/07 need raw text (ticker extraction + VADER); "
               "03/05/08/09/10 need only the aggregates. Splitting keeps the heavy "
               "text stage on the raw machine.",
        "technical detail": "export() publishes the aggregates producer-side; "
               "hydrate() copies them into data/processed on the consumer so the "
               "UNCHANGED notebooks find them where they already read - no notebook "
               "path edits.",
        "outcome": "No notebook surgery; clean machine boundary.",
        "date": TODAY,
    },
    {
        "decision": "Live append is text-free; weighted merge never revises history",
        "category": "Live data",
        "why": "New live posts must update the committable data without storing raw "
               "and without changing past values (live parity; snapshots never revised).",
        "technical detail": "api_calls/append_live_to_gic.py aggregates new posts "
               "(reusing build_mentions + sentiment) and merges deltas: mention_count "
               "adds; avg_sentiment and net_bullish recombine weighted by n_posts "
               "(combined = (a_old*n_old + a_new*n_new)/(n_old+n_new)). Proven identical "
               "to one-shot aggregation on real data (counts exact, sentiment to 1e-9).",
        "outcome": "History accumulates, never rewritten; the append keeps no text.",
        "date": TODAY,
    },
    {
        "decision": "First-seen-wins live dedup via local ledger + frozen LIVE_START",
        "category": "Live data",
        "why": "Re-running a fetch must not double-count, and live days must not "
               "collide with the committed historical block built from the dumps.",
        "technical detail": "data/reference/gic_live_meta.json (gitignored) holds the "
               "seen post-ids (capped 300k) and a frozen LIVE_START = newest committed "
               "day + 1. A candidate is dropped if dated before LIVE_START or if its id "
               "is already seen. The ledger stays LOCAL - post ids are mildly identifying "
               "so they are never committed.",
        "outcome": "Idempotent appends; no historical/live overlap; ids stay private.",
        "date": TODAY,
    },
    {
        "decision": "fetch_all auto-routes the append; run_daily gains --consumer mode",
        "category": "Orchestration",
        "why": "One command should do the right thing on each machine.",
        "technical detail": "fetch_all.py appends to posts.parquet when it exists, "
               "else folds into GIC_RAW_DATA (also forced by --gic). run_daily.py "
               "--consumer does fetch -> append_live_to_gic -> hydrate -> run the "
               "consumer notebooks 08/09/10 -> snapshot.",
        "outcome": "Work laptop: `python api_calls/fetch_all.py` then "
               "`python run_daily.py --consumer`.",
        "date": TODAY,
    },
]


def main():
    if not os.path.exists(XLSX):
        sys.exit(f"not found: {XLSX}")

    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active

    # ---- find the header row (row 1) and map header text -> column index
    header = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is not None:
            header[str(value).strip().lower()] = col
    print("columns found:", list(header.keys()))

    # which existing 'decision' texts are already there (skip duplicates)
    decision_col = header.get("decision", 1)
    existing = set()
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=decision_col).value
        if v:
            existing.add(str(v).strip())

    added = 0
    for d in DECISIONS:
        if d["decision"] in existing:
            print("skip (already present):", d["decision"][:50])
            continue
        target_row = ws.max_row + 1
        for key, value in d.items():
            col = header.get(key)
            if col is None:                 # header not found - append at the end
                col = ws.max_column + 1
                header[key] = col
                ws.cell(row=1, column=col).value = key
            ws.cell(row=target_row, column=col).value = value
        added += 1
        print("added:", d["decision"][:50])

    if added:
        wb.save(XLSX)
        print(f"\nsaved {added} new decision(s) -> {XLSX}")
    else:
        print("\nnothing to add - all decisions already logged.")


if __name__ == "__main__":
    main()
